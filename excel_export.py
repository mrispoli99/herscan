"""Build the annual-plan workbook for any territory/tech set (with booked-event overlay)."""
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

F='Arial'; NAVY='1F3864'; BLUE='2E5496'; HEAD='305496'
AMBER='FFE699'; ORANGE='F8CBAD'; GREEN='E2EFDA'; GREY='808080'; MON1='F2F2F2'
BOOKED='BDD7EE'; DEFER='FCE4D6'
_thin=Side(style='thin',color='D0D0D0'); BORDER=Border(_thin,_thin,_thin,_thin)

def _c(ws,r,c,v,bold=False,sz=10,color='000000',bg=None,align='left',wrap=False,bd=True):
    x=ws.cell(r,c,v); x.font=Font(name=F,bold=bold,size=sz,color=color)
    x.alignment=Alignment(horizontal=align,vertical='center',wrap_text=wrap)
    if bg: x.fill=PatternFill('solid',start_color=bg)
    if bd: x.border=BORDER
    return x

def build_workbook(path, asg, techs, grids, wdays, config, mets, df, names_map=None, weeks=52,
                   booked=None, booked_recon=None, extra_deferred=None):
    names=[t['name'] for t in techs]
    NAME = names_map or dict(zip(df['zip'].astype(str), df.get('Event Naming', df['zip'].astype(str))))
    drive_by={n:dict(zip(asg[asg.tech==n].zip, asg[asg.tech==n].drive)) for n in names}
    far_by={n:set(asg[(asg.tech==n)&(asg.far==1)].zip) for n in names}
    booked = booked or {}          # {(tech, week, day): {'date':'7/14','wkday':'Tue','label':'Aurora','oot':False}}
    start=dt.date.fromisoformat(config.get('start_date','2026-07-06'))
    wb=Workbook()

    # ---- Summary ----
    ws=wb.active; ws.title='Plan Summary'; ws.sheet_view.showGridLines=False
    ws.merge_cells('A1:F1'); _c(ws,1,1,config.get('title','Territory Annual Plan'),True,18,'FFFFFF',NAVY,bd=False)
    ws.merge_cells('A2:F2'); _c(ws,2,1,f"{weeks}-week calendar · {len(techs)} techs · starts week of {start.strftime('%-m/%-d/%y')}",
                               False,11,'FFFFFF',BLUE,bd=False); ws.row_dimensions[1].height=28
    r=4; _c(ws,r,1,'KEY NUMBERS',True,11,'FFFFFF',HEAD,bd=False); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
    tot=int(asg.visits_yr.sum())
    for a,b in [('Events / year',f'{tot}  (~{tot/weeks:.1f}/week)'),
                ('Territory ZIPs served',str(asg.zip.nunique())),
                ('Booked events folded in',str(len(booked))),
                ('Rotation',f"{config.get('rotation_weeks',8)}-week cycle within a {weeks}-week calendar")]:
        _c(ws,r,1,a,True,10,bd=False); ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6); _c(ws,r,2,b,bd=False); r+=1
    r+=1; _c(ws,r,1,'TECH LOADS',True,11,'FFFFFF',HEAD,bd=False); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
    hdr=['Tech','Home ZIP','Days/wk','Events/yr','Median drive','% long days (>75m)']
    for i,h in enumerate(hdr): _c(ws,r,1+i,h,True,10,'FFFFFF',BLUE,'center')
    r+=1
    for t in techs:
        m=mets[t['name']]
        vals=[t['name'],t.get('home_zip',''),t['days_per_week'],m['events'],f"{m['median']} min",f"{m['pct_long']:.0f}%"]
        for i,v in enumerate(vals): _c(ws,r,1+i,v,align='center' if i else 'left')
        r+=1
    r+=1; _c(ws,r,1,'COLOR KEY',True,11,'FFFFFF',HEAD,bd=False); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
    for lbl,bg in [('Booked event (locked to its date)',BOOKED),('Local day (≤45 min)',GREEN),
                   ('Longer day (71–90)',AMBER),('Far / overnight ✈ (>90)',ORANGE)]:
        _c(ws,r,1,lbl,bg=bg,align='center'); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2); r+=1
    notes=config.get('notes',[])
    if notes:
        r+=1; _c(ws,r,1,'NOTES',True,11,'FFFFFF',HEAD,bd=False); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6); r+=1
        for nt in notes:
            _c(ws,r,1,nt,sz=9,bd=False,wrap=True); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
            ws.row_dimensions[r].height=30; r+=1
    ws.column_dimensions['A'].width=24
    for col in 'BCDEF': ws.column_dimensions[col].width=18

    # ---- per-tech year tabs ----
    for t in techs:
        n=t['name']; days=wdays[n]; ncol=len(days)
        floating = days and not all(d in ['Mon','Tue','Wed','Thu','Fri','Sat','Sun'] for d in days)
        ws=wb.create_sheet(n[:28]); ws.sheet_view.showGridLines=False
        last=get_column_letter(2+ncol); ws.merge_cells(f'A1:{last}1')
        daydesc = 'floating days (booked events show their real date)' if floating else ', '.join(days)
        _c(ws,1,1,f"{n} — {weeks}-Week Calendar (home ZIP {t.get('home_zip','?')}, {t['days_per_week']} days/wk: {daydesc})",
           True,14,'FFFFFF',NAVY,bd=False); ws.row_dimensions[1].height=22
        _c(ws,2,1,'Week',True,10,'FFFFFF',HEAD,'center'); _c(ws,2,2,'Week of',True,10,'FFFFFF',HEAD,'center')
        for i,day in enumerate(days): _c(ws,2,3+i,day,True,10,'FFFFFF',HEAD,'center')
        ws.freeze_panes='A3'
        for w in range(weeks):
            rr=3+w; monday=start+dt.timedelta(weeks=w); msh=MON1 if monday.month%2==0 else 'FFFFFF'
            _c(ws,rr,1,w+1,True,9,'FFFFFF',BLUE,'center'); _c(ws,rr,2,monday.strftime('%-m/%-d/%y'),sz=9,color='404040',bg=msh,align='center')
            ws.row_dimensions[rr].height=32
            for i,day in enumerate(days):
                z=grids[n][w].get(day); key=(n,w,day)
                if not z: _c(ws,rr,3+i,'—',color=GREY,align='center'); continue
                if key in booked:                         # locked real booking
                    bk=booked[key]
                    if bk.get('oot'):
                        _c(ws,rr,3+i,f"{bk['label']}\n{bk.get('wkday','')} {bk['date']} · BOOKED (out of territory)",
                           sz=8,bold=True,bg=BOOKED,align='center',wrap=True)
                    else:
                        dm=bk.get('dm'); dm=drive_by[n].get(z) if dm is None else dm
                        dtxt=f"{int(dm)}m" if dm is not None else 'local'
                        far=' ⚠far' if (dm is not None and dm>=90) else ''
                        _c(ws,rr,3+i,f"{NAME.get(z,bk['label'])}{far}\n{z} · {bk.get('wkday','')} {bk['date']} · BOOKED · {dtxt}",
                           sz=8,bold=True,bg=BOOKED,align='center',wrap=True)
                    continue
                dm=drive_by[n].get(z,0); far=z in far_by[n]
                if far and dm>=115: bg,tag=ORANGE,' ✈ overnight'
                elif far: bg,tag=ORANGE,' ✈ long'
                elif dm>70: bg,tag=AMBER,' ○'
                elif dm<=45: bg,tag=GREEN,''
                else: bg,tag='FFFFFF',''
                _c(ws,rr,3+i,f"{NAME.get(z,z)}{tag}\n{z} · {int(dm)}m",sz=8,bg=bg,align='center',wrap=True)
        ws.column_dimensions['A'].width=6; ws.column_dimensions['B'].width=9
        for i in range(ncol): ws.column_dimensions[get_column_letter(3+i)].width=22

    # ---- Booked Events reconciliation ----
    if booked_recon:
        ws=wb.create_sheet('Booked Events'); ws.sheet_view.showGridLines=False
        _c(ws,1,1,'Already-booked events from the live schedule — how each was handled.',True,11,'FFFFFF',NAVY,bd=False)
        ws.merge_cells('A1:F1'); ws.row_dimensions[1].height=20
        for i,h in enumerate(['Date','City','Assigned tech','Disposition','Territory ZIP','Plan week']):
            _c(ws,2,1+i,h,True,10,'FFFFFF',HEAD,'center')
        ws.freeze_panes='A3'; r=3
        for x in booked_recon:
            fill = BOOKED if x['disp'].startswith('Fixed') else (DEFER if 'Defer' in x['disp'] else 'FFFFFF')
            for i,v in enumerate([x['date'],x['city'],x['tech'],x['disp'],x.get('zip') or '—',x.get('wk') or '—']):
                _c(ws,r,1+i,v,sz=9,bg=fill,align='center' if i in(0,4,5) else 'left')
            r+=1
        for i,wd in enumerate([12,26,16,34,12,9]): ws.column_dimensions[get_column_letter(1+i)].width=wd

    # ---- Territory ----
    ws=wb.create_sheet('Territory'); ws.sheet_view.showGridLines=False
    for i,h in enumerate(['ZIP','Location','Tech','Visits/yr','Drive (min)','Trip','Model rec/yr']): _c(ws,1,1+i,h,True,10,'FFFFFF',HEAD,'center')
    t2=asg.sort_values(['tech','visits_yr'],ascending=[True,False]); r=2
    for _,x in t2.iterrows():
        trip='Overnight' if x.drive>=115 else ('Long' if x.drive>90 else ('Longer' if x.drive>70 else 'Local'))
        for i,v in enumerate([x.zip,NAME.get(x.zip,x.zip),x.tech,int(x.visits_yr),int(x.drive),trip,int(x.rec)]):
            _c(ws,r,1+i,v,sz=9,align='center' if i in(0,3,4,6) else 'left')
        r+=1
    ws.freeze_panes='A2'
    for i,wd in enumerate([10,28,14,9,11,11,12]): ws.column_dimensions[get_column_letter(1+i)].width=wd

    # ---- Deferred (trimmed ZIPs + booked events handed off) ----
    served=set(asg.zip); d=df[~df['zip'].astype(str).isin(served)].copy()
    ws=wb.create_sheet('Deferred'); ws.sheet_view.showGridLines=False
    _c(ws,1,1,'Deferred — booked events handed off (other techs / out-of-territory) + ZIPs beyond current reach.',True,10,'FFFFFF',NAVY,bd=False)
    ws.merge_cells('A1:D1')
    for i,h in enumerate(['ZIP / Date','Location','Model rec/yr','Reason']): _c(ws,2,1+i,h,True,10,'FFFFFF',HEAD,'center')
    r=3
    for x in (extra_deferred or []):
        for i,v in enumerate([x.get('key',''),x.get('name',''),x.get('rec',''),x.get('reason','')]):
            _c(ws,r,1+i,v,sz=9,bg=DEFER,align='center' if i in(0,2) else 'left')
        r+=1
    d=d.sort_values('recommended_events_per_year',ascending=False)
    for _,x in d.iterrows():
        for i,v in enumerate([str(x['zip']),x.get('Event Naming',x['zip']),int(x['recommended_events_per_year']),x.get('scheduling_status','trimmed for capacity')]):
            _c(ws,r,1+i,v,sz=9,align='center' if i in(0,2) else 'left')
        r+=1
    for i,wd in enumerate([14,30,12,30]): ws.column_dimensions[get_column_letter(1+i)].width=wd

    wb.save(path); return path
